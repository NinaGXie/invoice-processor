#!/usr/bin/env python3
"""
Invoice Processor - Extract invoice data and create Excel summary
"""

import os
import re
import sys
from datetime import datetime
from pathlib import Path
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from pdf2image import convert_from_path
import pytesseract
from PIL import Image


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF file using pdfplumber, fallback to OCR if no text"""
    try:
        # First try pdfplumber for text-based PDFs
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

            # If we got meaningful text, return it
            if text.strip() and len(text.strip()) > 50:
                return text

            # Otherwise, try OCR
            print(f"    No text found, attempting OCR...")
            return extract_text_with_ocr(pdf_path)

    except Exception as e:
        # If pdfplumber fails, try OCR
        try:
            return extract_text_with_ocr(pdf_path)
        except Exception as ocr_error:
            return f"Error reading PDF: {str(e)}, OCR error: {str(ocr_error)}"


def extract_text_with_ocr(pdf_path):
    """Extract text from PDF using OCR (for image-based PDFs)"""
    try:
        # Check if tesseract is available
        try:
            pytesseract.get_tesseract_version()
        except:
            return "OCR not available: Tesseract not installed. Install with: brew install tesseract tesseract-lang"

        # Convert PDF to images
        images = convert_from_path(pdf_path, dpi=300)

        text = ""
        for image in images:
            # Use Chinese + English OCR
            page_text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            text += page_text + "\n"

        return text
    except Exception as e:
        return f"OCR Error: {str(e)}"


def extract_invoice_data(text, filename):
    """Extract amount, date, and destination from invoice text"""
    data = {
        'filename': filename,
        'amount': None,
        'date': None,
        'destination': None,
        'route': None,
        'full_text': text  # Store full text for categorization
    }

    # Extract amount (Chinese and English patterns)
    # Prioritize 价税合计 (tax-inclusive total)
    amount_patterns = [
        # Chinese patterns - prioritize 价税合计
        r'价税合计[\s:：]*[¥￥]?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # 价税合计: ¥1234.56
        r'价税合计.*?[¥￥]?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # 价税合计 1234.56
        r'(?:合计金额|总金额)[\s:：]*[¥￥]?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # 合计金额: ¥1234.56
        r'(?:小写|金额)[\s:：]*[¥￥]?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # 小写: ¥1234.56
        r'[¥￥]\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # ¥1,234.56
        # English patterns
        r'(?:Total|Amount|Sum)[\s:]*\$?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # Total: $1234.56
        r'\$\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',  # $1,234.56
        r'(\d+(?:,\d{3})*(?:\.\d{2})?)\s*(?:USD|EUR|GBP|CNY|RMB)',  # 1234.56 USD
    ]

    for pattern in amount_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount = match.group(1).replace(',', '')
            # Filter out unreasonably small amounts (likely not the total)
            if float(amount) >= 0.01:
                data['amount'] = amount
                break

    # Extract date (Chinese and English patterns)
    # For flight tickets, prioritize date in 备注 section (actual flight date)
    date_patterns = [
        # Flight ticket specific - date in 备注 section (e.g., 携程订单:xxx,2026/1/22)
        r'订单[:\s]*\d+,(\d{4}[/-]\d{1,2}[/-]\d{1,2})',  # 订单号,日期 format
        r',(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s+[\u4e00-\u9fa5]',  # ,日期 followed by Chinese characters
        # Chinese patterns
        r'(\d{4}年\d{1,2}月\d{1,2}日)',  # 2024年12月31日
        r'(?:开票日期|日期|时间)[\s:：]*(\d{4}年\d{1,2}月\d{1,2}日)',
        r'(?:开票日期|日期|时间)[\s:：]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
        # English patterns
        r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',  # 2024-12-31
        r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',  # 12/31/2024 or 12-31-2024
        r'((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4})',  # January 31, 2024
    ]

    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['date'] = match.group(1)
            break

    # Extract destination (Chinese and English patterns)
    # For Chinese invoices, we want the seller (销售方), not buyer (购买方)
    destination_patterns = [
        # Chinese patterns - prioritize seller, be more precise
        r'销\s+名称[\s:：]+([^统一社会\n]+)',  # 销 名称：
        r'售\s+名称[\s:：]+([^统一社会\n]+)',  # 售 名称
        r'销售?方[\s\S]{0,30}?名称[\s:：]*([^\n统一社会]{3,50}?)(?=\s*统一社会|$|\n)',  # 销售方 名称
        r'(?:收款方|付款方)[\s:：]+([^\n]{3,40})',
        r'(中国铁路|铁路客票)',  # For train tickets
        r'名\s*称[\s:：]+([^统一社会\n]{3,50})',  # Generic 名称
        r'(?:开票单位|开票方)[\s:：]+([^\n]{3,40})',
        # Ride-sharing specific
        r'(滴滴出行|曹操出行|高德打车|首汽约车)',
        # Hotel specific
        r'(.*?酒店|.*?宾馆|.*?旅馆)',
        # English patterns
        r'(?:Vendor|Company|From|Merchant)[\s:]+([A-Z][A-Za-z\s&,\.]+?)(?:\n|$)',
        r'(?:Bill To|Billed To)[\s:]+([A-Z][A-Za-z\s&,\.]+?)(?:\n|$)',
    ]

    for pattern in destination_patterns:
        match = re.search(pattern, text, re.MULTILINE | re.DOTALL)
        if match:
            dest = match.group(1).strip()
            # Clean up common noise
            dest = re.sub(r'[\s:：]+$', '', dest)
            dest = re.sub(r'\s+', '', dest)  # Remove extra spaces
            dest = re.sub(r'方$', '', dest)  # Remove trailing 方
            if len(dest) > 2 and not dest.isdigit():  # Avoid numbers and very short strings
                data['destination'] = dest
                break

    # Extract route for train/flight tickets (departure -> arrival)
    route_patterns = [
        # Train ticket: 北京南 G27 苏州北 (station, train number, station)
        r'([\u4e00-\u9fa5]{2,6})\s+[GDCKZTgdckzt]\d+\s+([\u4e00-\u9fa5]{2,6})\s*站',
        r'([\u4e00-\u9fa5]{2,6})\s+[GDCKZTgdckzt]\d+\s+([\u4e00-\u9fa5]{2,6})',
        # Flight: 上海(SHA) 北京(PEK)
        r'([^\s（(]{2,6})[（(][A-Z]{3}[)）]\s+([^\s（(]{2,6})[（(][A-Z]{3}[)）]',
        # Generic: 上海 -> 北京
        r'([\u4e00-\u9fa5]{2,6})\s*[-→至到]\s*([\u4e00-\u9fa5]{2,6})',
    ]
    for pattern in route_patterns:
        match = re.search(pattern, text)
        if match:
            dep = match.group(1).strip()
            arr = match.group(2).strip()
            if len(dep) >= 2 and len(arr) >= 2 and dep != arr:
                data['route'] = f"{dep}->{arr}"
                break

    return data


def categorize_invoice(destination, filename, full_text=''):
    """Categorize invoice type based on destination, filename, and full text"""
    text = (destination or '') + filename + (full_text or '')

    # 火车票 - Train tickets (check first for specificity)
    if any(keyword in text for keyword in ['火车', '铁路', '12306', '中国铁路', '动车', '高铁', '电子客票', '客票']):
        return '火车票'

    # 飞机票 - Flight tickets (check for specific keywords)
    elif any(keyword in text for keyword in ['飞机', '航空', '机票', '航班', 'airline', 'flight', 'airways', '代订机票', '退改费用', '经济舱', '商务舱']):
        return '飞机票'

    # 住宿 - Accommodation (exclude if it's travel agency selling flights)
    elif any(keyword in text for keyword in ['酒店', '住宿', '入住', '宾馆', 'hotel', 'inn', '民宿', 'airbnb']) and '机票' not in text:
        return '住宿'

    # 打车 - Taxi/Ride-sharing (check for specific companies and services)
    elif any(keyword in text for keyword in ['打车', '网约车', '滴滴', '曹操', '高德', '优步', 'uber', 'didi', '嘀嗒', '首汽', 'taxi', '出租车', '用车', '爱特博旅运', '吉利优行']):
        return '打车'

    # 餐饮 - Dining
    elif any(keyword in text for keyword in ['餐饮', '饭店', '餐厅', '美食', '食品', '咖啡', 'restaurant', 'cafe', 'coffee', '茶', '小吃', '快餐', '火锅', '烧烤']):
        return '餐饮'

    # 办公用品 - Office supplies
    elif any(keyword in text for keyword in ['文具', '办公', '打印', '复印', '纸张', 'office', 'stationery', '京东', 'jd', '淘宝', 'taobao', '天猫', 'tmall']):
        return '办公用品'

    # 快递物流 - Shipping
    elif any(keyword in text for keyword in ['快递', '物流', '邮政', '顺丰', '圆通', '中通', '韵达', 'express', 'shipping', 'ems']):
        return '快递'

    # 通讯 - Communication
    elif any(keyword in text for keyword in ['通信', '电信', '移动', '联通', '话费', 'telecom', 'mobile', '流量']):
        return '通讯'

    else:
        return '其他'


def extract_month_day(date_str):
    """Extract month and day from date string"""
    if not date_str:
        return None, None

    # Try to extract month and day from Chinese format: 2026年02月10日
    match = re.search(r'(\d{1,2})月(\d{1,2})日', date_str)
    if match:
        return match.group(1), match.group(2)

    # Try YYYY/M/D or YYYY-M-D format: 2026/1/22
    match = re.search(r'\d{4}[-/](\d{1,2})[-/](\d{1,2})', date_str)
    if match:
        return match.group(1), match.group(2)

    # Try other formats M/D or M-D
    match = re.search(r'^(\d{1,2})[-/](\d{1,2})', date_str)
    if match:
        return match.group(1), match.group(2)

    return None, None


def create_excel(data, output_path):
    """Create Excel file with invoice summary"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Summary"

    # Headers
    headers = ["序号", "月", "日", "报销内容", "报销金额", "发票类型", "事项", "项目名称"]
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows
    for idx, invoice in enumerate(data, 1):
        month, day = extract_month_day(invoice['date'])
        category = categorize_invoice(invoice['destination'], invoice['filename'], invoice.get('full_text', ''))
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)           # 序号
        ws.cell(row=row, column=2, value=int(month) if month else None)  # 月
        ws.cell(row=row, column=3, value=int(day) if day else None)      # 日
        ws.cell(row=row, column=4, value=category)      # 报销内容 (e.g. 火车票)
        # 报销金额 as number
        try:
            ws.cell(row=row, column=5, value=float(invoice['amount']) if invoice['amount'] else None)
        except (ValueError, TypeError):
            ws.cell(row=row, column=5, value=invoice['amount'])
        ws.cell(row=row, column=6, value=None)          # 发票类型 (blank, dropdown)
        # 事项: route for train/flight, otherwise blank
        route = invoice.get('route')
        if category in ('火车票', '飞机票') and route:
            ws.cell(row=row, column=7, value=route)
        else:
            ws.cell(row=row, column=7, value=None)
        ws.cell(row=row, column=8, value=None)          # 项目名称

    # 合计 row — pre-calculated sum (works on mobile apps that don't evaluate formulas)
    total_row = len(data) + 2
    total_label = ws.cell(row=total_row, column=1, value="合计")
    total_label.font = Font(bold=True)
    total_amount = 0
    for invoice in data:
        try:
            total_amount += float(invoice['amount']) if invoice['amount'] else 0
        except (ValueError, TypeError):
            pass
    sum_cell = ws.cell(row=total_row, column=5, value=round(total_amount, 2))
    sum_cell.font = Font(bold=True)

    # Dropdown for 发票类型 (column F)
    dv = DataValidation(
        type="list",
        formula1='"纸质普通发票,纸质专用发票,电子普通发票,电子专用发票"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.sqref = f"F2:F{total_row - 1}"

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20

    wb.save(output_path)


def main():
    print("=" * 60)
    print("Invoice Processor")
    print("=" * 60)

    # Get folder path
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = input("\nEnter the path to your invoice folder: ").strip()

    if not os.path.exists(folder):
        print(f"Error: Folder '{folder}' does not exist")
        sys.exit(1)

    # Find all PDF files
    pdf_files = list(Path(folder).glob("*.pdf"))

    if not pdf_files:
        print(f"\nNo PDF files found in '{folder}'")
        sys.exit(1)

    print(f"\nFound {len(pdf_files)} PDF file(s)")
    print("\nProcessing invoices...")
    print("-" * 60)

    # Process each invoice
    invoice_data = []
    for i, pdf_file in enumerate(pdf_files, 1):
        print(f"[{i}/{len(pdf_files)}] {pdf_file.name}...", end=" ")

        text = extract_text_from_pdf(pdf_file)
        data = extract_invoice_data(text, pdf_file.name)
        invoice_data.append(data)

        print("✓")

    # Create Excel file
    output_path = os.path.join(folder, f"invoice_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    create_excel(invoice_data, output_path)

    print("-" * 60)
    print(f"\n✓ Successfully processed {len(pdf_files)} invoice(s)")
    print(f"✓ Excel file created: {output_path}")
    print("\nSummary:")
    for invoice in invoice_data:
        print(f"  • {invoice['filename']}")
        print(f"    Amount: {invoice['amount'] or 'Not found'}")
        print(f"    Date: {invoice['date'] or 'Not found'}")
        print(f"    Destination: {invoice['destination'] or 'Not found'}")
        print()


if __name__ == "__main__":
    main()
